#!/usr/bin/env python3
"""export_wcheck.py - 証拠台帳とWチェック結果を1シートに統合してExcel出力する

設計方針（重要）: Wチェックの状態は固定語彙（問題なし/懸念あり/問題あり/未実施）のみを使い、
Claudeが自由記述で判定を書き込むことはしない。理由文も、各チェックが実際に出力した
メッセージをそのまま転記する（新たな文章をここで作らない）。

自由記述を許すと、詳細な理由文ほど「もっともらしく見える」ため、
中身が伴わない記述で検証を通過できてしまう（過去の性悪説監査で判明した抜け穴と同型）。
固定語彙＋機械生成の理由文に限定することで、この経路を塞ぐ。

統合内容:
  - 証拠台帳の必須7項目（eid/source_id/claim_type/anchor_head/anchor_tail/
    basis_date/confidence/tags）＋任意項目（value/subject/note/char_start/char_end 等）
  - Wチェック結果 C-1/C-2/C-3（判定＋理由）
  任意項目が空の証拠は「出力対象外」と明記する（空欄と未取得を区別するため）。

Wチェック3層の対応（言葉の定義）:
  C-1 レポート本文 → レポート抜粋/要約 の整合 … claims_review.md の判定列（GC＝独立レビュー）
  C-2 レポート抜粋/要約 → ソース の整合       … reverify_log.jsonl の outcome（原典再アクセス）
  C-3 文脈整合（台帳の引用がソース内の限定・前提・注釈を取りこぼしていないか）
                                            … check_a の A15/A16/A18/A19/A20/A21
                                              （A17はソース種別の妥当性で別軸のため含めない）

各ヘッダセルには、その項目の定義説明をExcelのセルメモ（コメント）として付与する。

usage: python3 scripts/export_wcheck.py <case_dir> [--out path.xlsx] [--no-xlsx]
出力: <case_dir>/checks/wcheck_summary.jsonl（常に生成、機械ファイル）
      <case_dir>/checks/wcheck_summary.xlsx（--no-xlsx で省略可）
"""
from __future__ import annotations

import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from harness_lib import (  # noqa: E402
    case_path, read_text, load_ledger, load_reverify_log, load_sources,
)
import check_a_source_ledger as check_a  # noqa: E402

# 固定語彙。この4値以外を使わない。
NO_ISSUE = "問題なし"
CONCERN = "懸念あり"
ISSUE = "問題あり"
NOT_DONE = "未実施"

# 任意項目が空のときに入れる明示マーカー（空欄＝未処理と区別する）
NA = "出力対象外"

# C-3（文脈整合）は「台帳の引用が原文内の限定・前提・注釈を取りこぼしていないか」を見る。
# A17（ソース種別の妥当性）は文脈整合ではなく別軸のため、checker.md の正式定義どおり除外する。
# A27は局所窓(A18)と冒頭ゾーン(A19)の中間を埋める拡張文脈スキャン（当該ページ＋次ページ／WEBは全体）。
C3_CODES = {"A15", "A16", "A18", "A19", "A20", "A21", "A27"}

C2_OUTCOME_MAP = {
    "match": NO_ISSUE,
    "context_reversed": CONCERN,
    "hallucination": ISSUE,
    "unreachable": NOT_DONE,
}
C1_VERDICT_MAP = {
    "支持": NO_ISSUE,
    "判断不能": CONCERN,
    "部分的": CONCERN,
    "不支持": ISSUE,
}

# ---- 各項目の定義（ヘッダのセルメモに入れる） ---------------------------------
# key は内部列キー。value はメモに表示する定義文。
FIELD_DEFS = {
    "eid": (
        "証拠ID。原文の完全一致部分文字列に紐づく一意な識別子。全事実記述はこのIDを持つ。\n"
        "書式: E-xxxx（4桁連番）"
    ),
    "source_id": (
        "この証拠が属するソースのID。sources.json のエントリに対応する。\n"
        "書式: S-xxx（3桁連番）"
    ),
    "src_link": (
        "ソースの到達URL（sources.json の url_or_path）。PDFは該当ページを #page=n で付す。\n"
        "URLが無くローカル保存のみの場合は local_path を表示。どちらも無ければ『出力対象外』。"
    ),
    "src_page": (
        "証拠の引用箇所が存在するソース内ページ番号。PDF抽出時の<<<PAGE n>>>マーカーから、\n"
        "char_start直前のページを機械的に特定。マーカーが無い形式（HTML等）は『出力対象外』。"
    ),
    "subject": (
        "【任意】主張の主体（セグメント名・子会社名・製品名等）。\n"
        "セグメント別数値を扱う場合は必須。空なら『出力対象外』。"
    ),
    "fact": (
        "その証拠を自分の言葉で正規化した記述（単位と期を明示）。原文の正規化であり、\n"
        "レポート本文に採用したか否かとは別（台帳には未採用の証拠も残る）。"
    ),
    "claim_type": (
        "証拠の種別。取りうる値:\n"
        "  fact＝実績値 / negative＝非公表の確認 / forecast＝会社予想 / opinion＝第三者見解"
    ),
    "confidence": (
        "証拠自体の確からしさ。ソース階層と裏取り状況に基づく。取りうる値:\n"
        "  高 / 中 / 低"
    ),
    "tags": (
        "調査カテゴリ（10区分から複数可）。取りうる値:\n"
        "  ビジネスモデル / 財務.実績 / 財務.株価指標 / 財務.健全性 / 市場環境 /\n"
        "  競争環境 / 技術 / ガバナンス / リスク / 株主還元"
    ),
    "anchor_head": (
        "引用範囲の開始を示す識別語。原文冒頭を40字以内で写した位置索引。\n"
        "原文の再現ではなく位置の索引として使う。"
    ),
    "anchor_tail": "引用範囲の終了を示す識別語。原文末尾を40字以内で写した位置索引。",
    "basis_date": (
        "情報の基準日（決算期末や発表日）。取得日ではない。\n"
        "書式: YYYY-MM-DD"
    ),
    "value": (
        "【任意】数値の構造化。フィールド:\n"
        "  number＝数値 / unit＝単位 / period＝対象期 / basis_scope＝集計範囲(連結/単体・累計/単四半期)\n"
        "数値証拠でない場合は『出力対象外』。"
    ),
    "note": (
        "【任意】会計基準・連単・調整後指標の定義などの前提。\n"
        "株価指標では基準（会社予想/コンセンサス/実績）と対象期を書く。無ければ『出力対象外』。"
    ),
    "char_range": (
        "【任意】raw_text内の引用位置 [char_start:char_end]。build_ledger.pyがanchorから自動解決。\n"
        "未解決なら『出力対象外』。"
    ),
    # ---- Wチェック3層。判定は固定4語彙、理由は各チェックが出した文の転記のみ ----
    "c1_status": (
        "C-1判定：レポート本文 → レポート抜粋/要約（台帳のfact）の整合。\n"
        "レポート本文が、その根拠とした抜粋/要約を正しく使っているか。\n"
        "出所: claims_review.md（GC＝独立レビュー）の判定を固定語彙に写す。\n"
        "取りうる値: 問題なし＝支持 / 懸念あり＝部分的・判断不能 / 問題あり＝不支持 / 未実施＝GC未記入"
    ),
    "c1_reason": (
        "C-1の理由。claims_review.md（GCレビュー）のコメント列を転記。\n"
        "このスクリプトで新規作文はしない（結果の言い換えを禁じる）。"
    ),
    "c2_status": (
        "C-2判定：レポート抜粋/要約 → ソース（原典）の整合。\n"
        "抜粋/要約が、原典に再アクセスして現物と一致するか。\n"
        "出所: reverify_log.jsonl の outcome を固定語彙に写す。\n"
        "取りうる値: 問題なし＝match / 懸念あり＝context_reversed / 問題あり＝hallucination /\n"
        "  未実施＝unreachable（原典にアクセスできず未検証）または再アクセス記録が無い。\n"
        "※未実施は『検証していない』の明示。Critical証拠が未実施だと check_a A23 が別途FAILにする。"
    ),
    "c2_reason": (
        "C-2の理由。原典再アクセスを行った者が reverify_log.jsonl の detail に記録した\n"
        "確認内容を転記。このスクリプトで新規作文はしない。"
    ),
    "c3_status": (
        "C-3判定：文脈整合。台帳の引用が、ソース（原文）内にある限定・前提・注釈・\n"
        "反対文脈を取りこぼしていないか。レポート本文には現れないが原文側で意味を狭める/\n"
        "反転させる条件の脱落も対象に含む。\n"
        "出所: check_a を再実行し A15/A16/A18/A19/A20/A21/A27 の結果を集約。\n"
        "  A15＝単位の裏付け / A16＝期の一致 / A18＝直前直後（前後100字）の留保・条件の脱落 /\n"
        "  A19＝文書冒頭にある前提ゾーンからの乖離 / A20＝連結単体・累計単四半期の一致 /\n"
        "  A21＝複数期要約表からの引用（粒度不明） /\n"
        "  A27＝当該ページ全体（PDFは＋次ページ／WEBはソース全体）の留保・条件の脱落\n"
        "取りうる値: 問題なし＝指摘なし / 懸念あり＝WARN / 問題あり＝FAIL"
    ),
    "c3_reason": (
        "C-3の理由。check_a が実際に出力した指摘メッセージを転記。\n"
        "指摘が無い場合は『A15〜A21・A27で指摘なし』。"
    ),
}

# 出力する列の順序と表示ヘッダ名（内部キー -> 表示名）
COLUMNS = [
    ("eid", "EID"),
    ("source_id", "Source"),
    ("src_link", "Sourceリンク(ページ付)"),
    ("src_page", "掲載ページ"),
    ("subject", "Subject"),
    ("fact", "Fact"),
    ("claim_type", "種別"),
    ("confidence", "確信度"),
    ("tags", "タグ"),
    ("anchor_head", "anchor_head"),
    ("anchor_tail", "anchor_tail"),
    ("basis_date", "基準日"),
    ("value", "value(数値構造)"),
    ("note", "note(前提)"),
    ("char_range", "char範囲"),
    ("c1_status", "C-1判定"),
    ("c1_reason", "C-1理由(GC)"),
    ("c2_status", "C-2判定"),
    ("c2_reason", "C-2理由(原典再アクセス)"),
    ("c3_status", "C-3判定"),
    ("c3_reason", "C-3理由(文脈整合)"),
]

PAGE_RE = re.compile(r"<<<PAGE (\d+)>>>")


def resolve_page(case_dir, ev):
    """証拠のchar_startから、raw_text内の直前<<<PAGE n>>>マーカーを逆探索してページ番号を返す。

    マーカーが存在しない形式（HTML等）や、位置未解決の場合は None（→出力対象外）。
    """
    cs = ev.get("char_start")
    sid = ev.get("source_id")
    if cs is None or not sid:
        return None
    txt = read_text(case_path(case_dir, "raw_text", f"{sid}.txt"))
    if not txt or "<<<PAGE" not in txt:
        return None
    last = None
    for m in PAGE_RE.finditer(txt):
        if m.start() <= cs:
            last = m.group(1)
        else:
            break
    return last


def build_src_link(src, page):
    """sources.json のエントリからリンク文字列を作る。PDFかつページ判明時は #page=n を付す。"""
    if not src:
        return NA
    url = src.get("url_or_path") or src.get("local_path") or ""
    if not url:
        return NA
    if page and url.lower().endswith(".pdf"):
        return f"{url}#page={page}"
    if page and ("edinet" in url.lower() or "/pdf" in url.lower()):
        return f"{url}#page={page}"
    return url


def fmt_optional(val):
    """任意項目を表示用に整形。空・None は『出力対象外』にする。"""
    if val is None:
        return NA
    if isinstance(val, dict):
        if not val:
            return NA
        return json.dumps(val, ensure_ascii=False)
    if isinstance(val, (list, tuple)):
        return ",".join(str(x) for x in val) if val else NA
    s = str(val).strip()
    return s if s else NA


def collect_c3(case_dir):
    res = check_a.run(case_dir)
    by_eid = {}
    for f in res.findings:
        if f.code not in C3_CODES:
            continue
        eid = f.where
        entry = by_eid.setdefault(eid, {"level": None, "reasons": []})
        entry["reasons"].append(f.line())
        if f.level == "FAIL":
            entry["level"] = "FAIL"
        elif f.level == "WARN" and entry["level"] != "FAIL":
            entry["level"] = "WARN"
    out = {}
    for eid, entry in by_eid.items():
        status = ISSUE if entry["level"] == "FAIL" else CONCERN
        out[eid] = (status, " / ".join(entry["reasons"]))
    return out


def collect_c2(case_dir):
    out = {}
    for row in load_reverify_log(case_dir):
        eid = row.get("eid")
        if not eid:
            continue
        status = C2_OUTCOME_MAP.get(row.get("outcome", ""))
        if status is None:
            continue
        out[eid] = (status, str(row.get("detail", "")))
    return out


ROW_RE = re.compile(r"^\|\s*(\d+)\s*\|(.*)\|\s*$")


def collect_c1(case_dir):
    text = read_text(case_path(case_dir, "checks", "claims_review.md"))
    if not text:
        return {}
    severity_order = {NO_ISSUE: 0, CONCERN: 1, ISSUE: 2}
    out = {}
    for line in text.splitlines():
        m = ROW_RE.match(line.strip())
        if not m or not re.match(r"^\d+$", m.group(1)):
            continue
        cells = [c.strip() for c in m.group(2).split("|")]
        if len(cells) < 4:
            continue
        # EIDは列位置に依存せずパターンで探す（表レイアウト変更に強くする）。
        # 1セルに複数E-IDが入る場合は全てに同じ判定を割り当てる。
        eids = []
        for c in cells:
            eids.extend(re.findall(r"E-\d{4}", c))
        if not eids:
            continue
        # 判定は固定語彙のいずれかに一致するセルを採る（無ければスキップ）。
        verdict = None
        for c in cells:
            if c in C1_VERDICT_MAP:
                verdict = c
                break
        status = C1_VERDICT_MAP.get(verdict)
        if status is None:
            continue
        comment = cells[-1] if cells else ""
        for eid in eids:
            if eid not in out or severity_order[status] > severity_order[out[eid][0]]:
                out[eid] = (status, comment)
    return out


def build_summary(case_dir):
    evidence = {r.get("eid"): r for _, r in load_ledger(case_dir) if r.get("eid")}
    sources, _ = load_sources(case_dir)
    c1 = collect_c1(case_dir)
    c2 = collect_c2(case_dir)
    c3 = collect_c3(case_dir)

    rows = []
    for eid, ev in sorted(evidence.items()):
        page = resolve_page(case_dir, ev)
        src = sources.get(ev.get("source_id"))
        s1, r1 = c1.get(eid, (NOT_DONE, ""))
        s2, r2 = c2.get(eid, (NOT_DONE, ""))
        s3, r3 = c3.get(eid, (NO_ISSUE, "A15〜A21・A27で指摘なし"))

        cs, ce = ev.get("char_start"), ev.get("char_end")
        char_range = f"[{cs}:{ce}]" if cs is not None and ce is not None else NA

        rows.append({
            "eid": eid,
            "source_id": ev.get("source_id", ""),
            "src_link": build_src_link(src, page),
            "src_page": page if page else NA,
            "subject": fmt_optional(ev.get("subject")),
            "fact": ev.get("fact", ""),
            "claim_type": ev.get("claim_type", ""),
            "confidence": ev.get("confidence", ""),
            "tags": ",".join(ev.get("tags", []) or []),
            "anchor_head": ev.get("anchor_head", ""),
            "anchor_tail": ev.get("anchor_tail", ""),
            "basis_date": ev.get("basis_date", ""),
            "value": fmt_optional(ev.get("value")),
            "note": fmt_optional(ev.get("note")),
            "char_range": char_range,
            "c1_status": s1, "c1_reason": fmt_optional(r1) if s1 != NOT_DONE else "",
            "c2_status": s2, "c2_reason": fmt_optional(r2) if s2 != NOT_DONE else "",
            "c3_status": s3, "c3_reason": r3,
        })
    return rows


def write_jsonl_summary(case_dir, rows):
    path = case_path(case_dir, "checks", "wcheck_summary.jsonl")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        for r in rows:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")
    return path


def export_xlsx(case_dir, rows, out_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
    except ImportError:
        print("openpyxl未導入のためExcel出力をスキップします。"
              "pip install openpyxl --break-system-packages")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wチェック統合"

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    # ヘッダ行を書き、各セルに定義メモ（コメント）を付与
    for col_idx, (key, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        definition = FIELD_DEFS.get(key)
        if definition:
            cmt = Comment(definition, "equity-research-harness")
            cmt.width = 320
            cmt.height = 160
            cell.comment = cmt

    status_fill = {
        ISSUE: PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        CONCERN: PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        NO_ISSUE: PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        NOT_DONE: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    na_font = Font(color="808080", italic=True)
    status_cols = {"c1_status", "c2_status", "c3_status"}

    for r in rows:
        excel_row = ws.max_row + 1
        for col_idx, (key, _label) in enumerate(COLUMNS, 1):
            val = r.get(key, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if key in status_cols:
                fill = status_fill.get(val)
                if fill:
                    cell.fill = fill
            elif val == NA:
                cell.font = na_font

    widths = {
        "EID": 8, "Source": 8, "Sourceリンク(ページ付)": 30, "掲載ページ": 8,
        "Subject": 14, "Fact": 30, "種別": 8, "確信度": 7, "タグ": 16,
        "anchor_head": 16, "anchor_tail": 16, "基準日": 11,
        "value(数値構造)": 22, "note(前提)": 20, "char範囲": 12,
        "C-1判定": 9, "C-1理由(GC)": 30, "C-2判定": 9,
        "C-2理由(原典再アクセス)": 30, "C-3判定": 9, "C-3理由(文脈整合)": 30,
    }
    for col_idx, (_key, label) in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(label, 14)
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    case_dir = sys.argv[1]
    no_xlsx = "--no-xlsx" in sys.argv
    out_path = None
    if "--out" in sys.argv:
        out_path = sys.argv[sys.argv.index("--out") + 1]
    elif not no_xlsx:
        out_path = case_path(case_dir, "checks", "wcheck_summary.xlsx")

    rows = build_summary(case_dir)
    jsonl_path = write_jsonl_summary(case_dir, rows)
    print(f"{len(rows)}件を集約 → {jsonl_path}")

    counts = {}
    for r in rows:
        for k in ("c1_status", "c2_status", "c3_status"):
            counts[(k, r[k])] = counts.get((k, r[k]), 0) + 1
    for dim in ("c1_status", "c2_status", "c3_status"):
        line = " / ".join(f"{v}:{counts.get((dim, v), 0)}"
                          for v in (NO_ISSUE, CONCERN, ISSUE, NOT_DONE))
        print(f"  {dim}: {line}")

    if out_path:
        path = export_xlsx(case_dir, rows, out_path)
        if path:
            print(f"→ Excel: {path}")


if __name__ == "__main__":
    main()
