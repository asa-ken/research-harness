#!/usr/bin/env python3
"""run_tests.py - 全チェックコードの網羅テスト

各コードについて「発火すべきケース」と「発火してはいけないケース」を対で持つ。
片方だけだと、検出できても誤検出だらけ（あるいはその逆）の状態を見逃すため。

usage: python3 tests/run_tests.py [-v]
終了コード: 全件成功で0
"""
from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(SKILL, "scripts"))

import check_a_source_ledger as check_a  # noqa: E402
import check_b_ledger_report as check_b  # noqa: E402
import check_l_search_log as check_l  # noqa: E402
import lint_status  # noqa: E402
from harness_lib import sha256_text  # noqa: E402

VERBOSE = "-v" in sys.argv

BASE_RAW = """当連結会計年度の売上高は1,234,567百万円となりました。
営業利益は150,000百万円となりました。
産業機器セグメントの売上高は700,000百万円であります。
産業機器セグメントの出荷台数は12.0百万台となりました。
前連結会計年度の産業機器セグメントの売上高は650,000百万円でありました。
当社は上位顧客への依存度が高く、業績が変動する可能性があります。
なお、セグメント別の設備投資額については記載しておりません。
次期の売上高は1,300,000百万円を見込んでおります。
"""

PRICE_RAW = """東証PRM 15:30 終値 3,132円
PER（会社予想） (連)11.79倍
PBR（実績） (連)0.99倍
"""


def ev(**kw):
    """テスト用の証拠。quote= で渡された場合は anchor_head/tail に自動変換する。

    新形式では原文を転記せず識別語で指すため、テストも同じ形に揃える。
    """
    base = {"eid": "E-0001", "source_id": "S-001", "claim_type": "fact",
            "anchor_head": "産業機器セグメントの売上高は",
            "anchor_tail": "であります。",
            "fact": "産業機器セグメント売上 700,000百万円",
            "basis_date": "2025-06-20", "confidence": "高", "tags": ["財務.実績"]}
    q = kw.pop("quote", None)
    if q is not None:
        # 先頭・末尾から識別語を切り出す（40字以内）
        base["anchor_head"] = q[:min(20, max(4, len(q)//3))]
        base["anchor_tail"] = q[-min(20, max(4, len(q)//3)):]
    base.update(kw)
    return base


def src(**kw):
    base = {"source_id": "S-001", "title": "有報", "publisher": "EDINET",
            "source_class": "T1", "published_at": "2025-06-20",
            "period_context": "25年3月期", "intended_use": "検証"}
    base.update(kw)
    return base


def build(tmp, raw=None, sources=None, ledger=None, report=None, figures=None):
    for sub in ("raw_text", "figures", "checks", "requests"):
        os.makedirs(os.path.join(tmp, sub), exist_ok=True)
    raw = raw if raw is not None else {"S-001": BASE_RAW}
    for sid, text in raw.items():
        with open(os.path.join(tmp, "raw_text", f"{sid}.txt"), "w", encoding="utf-8") as fh:
            fh.write(text)
    sources = sources if sources is not None else [src()]
    for s in sources:
        sid = s["source_id"]
        if sid in raw and "text_sha256" not in s:
            s["text_sha256"] = sha256_text(raw[sid])
    with open(os.path.join(tmp, "sources.json"), "w", encoding="utf-8") as fh:
        json.dump(sources, fh, ensure_ascii=False)
    with open(os.path.join(tmp, "ledger.jsonl"), "w", encoding="utf-8") as fh:
        for row in (ledger if ledger is not None else [ev()]):
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
    if report is not None:
        with open(os.path.join(tmp, "report.md"), "w", encoding="utf-8") as fh:
            fh.write(report)
    for name, body in (figures or {}).items():
        with open(os.path.join(tmp, "figures", f"{name}.md"), "w", encoding="utf-8") as fh:
            fh.write(body)
    with open(os.path.join(tmp, "requests", "needed_sources.md"), "w", encoding="utf-8") as fh:
        fh.write("# 依頼\n\n### R-001 [優先度:H]\n- 欲しいもの: テスト\n")
    return tmp


FIG_OK = {f"F-0{i}": ("# F\n\n## データ表\n| 項目 | 値 | 根拠 |\n|---|---|---|\n"
                      "| 売上 | 700000 | E-0001 |\n") for i in (1, 2, 3, 4, 5)}

RESULTS = []


def check(name, kind, fixture, expect=(), forbid=()):
    """kind: a=チェックA / b=チェックB / s=ステータスLint"""
    tmp = tempfile.mkdtemp()
    try:
        build(tmp, **fixture)
        if kind == "a":
            res = check_a.run(tmp)
        elif kind == "s":
            res = lint_status.run(tmp)
        else:
            res = check_b.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            mark = "PASS" if ok else "FAIL"
            print(f"[{mark}] {name}")
            if missing:
                print(f"       検出されるべきだが出なかった: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
            if VERBOSE:
                print(f"       実際: {sorted(codes)}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# =====================================================================
# チェックA
# =====================================================================
check("A01 引用が原文に無い（捏造）", "a",
      {"ledger": [ev(quote="当社は業界最大手であり圧倒的なシェアを持つ。")]},
      expect=["A01"])
check("A01 正常な引用は発火しない", "a", {}, forbid=["A01"])
check("A01 PDF由来の改行分断でも識別語で解決できる", "a",
      {"raw": {"S-001": "産業機器セグメントの\n売上高は700,000百万円\nであります。"}},
      forbid=["A01"])
check("A01 識別語が原文に無ければFAIL（改変・記憶からの記述）", "a",
      {"ledger": [ev(anchor_head="産業機器セグメントの売上高は７００",
                    anchor_tail="００百万円であります。")]},
      expect=["A01"])
check("A03 オフセット不一致", "a",
      {"ledger": [ev(char_start=0, char_end=10)]}, expect=["A03"])
check("A04 source_idが存在しない", "a",
      {"ledger": [ev(source_id="S-999")]}, expect=["A04"])
check("A05 原文が差し替えられている", "a",
      {"sources": [src(text_sha256="0" * 64)]}, expect=["A05"])
check("A06 登録ソースが未使用", "a",
      {"raw": {"S-001": BASE_RAW, "S-002": "別文書"},
       "sources": [src(), src(source_id="S-002")]}, expect=["A06"])
check("A06 unused_reason があれば黙る", "a",
      {"raw": {"S-001": BASE_RAW, "S-002": "別文書"},
       "sources": [src(), src(source_id="S-002", unused_reason="参考のみ")]},
      forbid=["A06"])
check("A07 必須フィールド欠落", "a",
      {"ledger": [{"eid": "E-0001", "source_id": "S-001", "quote": "産業機器"}]},
      expect=["A07"])
check("A07 E-ID重複", "a", {"ledger": [ev(), ev()]}, expect=["A07"])
check("A08 value.numberが引用に無い", "a",
      {"ledger": [ev(value={"number": 155000, "unit": "百万円", "period": "25年3月期"})]},
      expect=["A08"])
check("A08 正しい数値は発火しない", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期"})]},
      forbid=["A08"])
check("A09 低Tierなのにconfidence高", "a",
      {"sources": [src(source_class="T5")]}, expect=["A09"])
check("A10 基準日が古い", "a",
      {"ledger": [ev(basis_date="2020-01-01")]}, expect=["A10"])
check("A15 単位の取り違え（百万円→億円）", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "億円", "period": "25年3月期"})]},
      expect=["A15"])
check("A15 正しい単位は発火しない", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期"})]},
      forbid=["A15"])
check("A16 宣言済み対象期との食い違い", "a",
      {"ledger": [ev(quote="当連結会計年度の売上高は1,234,567百万円となりました。",
                     value={"number": 1234567, "unit": "百万円", "period": "24年3月期"})]},
      expect=["A16"])
check("A16 相対表現はWARN止まり", "a",
      {"ledger": [ev(quote="前連結会計年度の産業機器セグメントの売上高は650,000百万円でありました。",
                     value={"number": 650000, "unit": "百万円", "period": "24年3月期"})]},
      expect=["A16"])
check("A16 period_contextと一致すれば黙る", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期"})]},
      forbid=["A16"])

PRICE_SRC = [src(source_id="S-002", source_class="T4", period_context="")]
PRICE_EV = {"eid": "E-0001", "source_id": "S-002", "claim_type": "fact",
            "quote": "東証PRM 15:30 終値 3,132円", "fact": "終値3,132円",
            "value": {"number": 3132, "unit": "円", "period": "2026-08-21"},
            "basis_date": "2026-08-21", "confidence": "高",
            "tags": ["財務.株価指標"], "note": "東証終値"}
check("A11 株価にbasis_datetimeが無い", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC, "ledger": [dict(PRICE_EV)]},
      expect=["A11"])
check("A11 時刻付きなら黙る", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, basis_datetime="2026-08-21T15:30+09:00")]},
      forbid=["A11", "A13"])
check("A12 会社予想ベースの指標は採用基準外", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, quote="PER（会社予想） (連)11.79倍", fact="PER 11.79倍",
                       value={"number": 11.79, "unit": "倍", "period": "2026-08-21"},
                       basis_datetime="2026-08-21T15:30+09:00",
                       note="会社予想EPS(27年3月期)基準、株価は終値")]},
      expect=["A12"])
check("A12 コンセンサス基準なら通る", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, quote="PER（会社予想） (連)11.79倍", fact="PER 11.79倍",
                       value={"number": 11.79, "unit": "倍", "period": "2026-08-21"},
                       basis_datetime="2026-08-21T15:30+09:00",
                       note="コンセンサス予想EPS(27年3月期)基準、株価は終値")]},
      forbid=["A12"])
check("A13 場中の株価（引け前）", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, basis_datetime="2026-08-21T14:00+09:00")]},
      expect=["A13"])
check("A13 リアルタイム値の明示", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, basis_datetime="2026-08-21T15:30+09:00",
                       note="リアルタイム株価")]},
      expect=["A13"])
check("A14 土日の日付", "a",
      {"raw": {"S-002": PRICE_RAW}, "sources": PRICE_SRC,
       "ledger": [dict(PRICE_EV, basis_date="2026-08-22",
                       basis_datetime="2026-08-22T15:30+09:00")]},
      expect=["A14"])

# =====================================================================
# キーワード1語での回避を防ぐ（substantive_ack）
# =====================================================================
check("A15 単語1つの'単位'ではFAILを黙らせられない", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "億円", "period": "25年3月期"},
                    note="単位")]},
      expect=["A15"])
check("A15 実質的な換算根拠なら通る", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "億円", "period": "25年3月期"},
                    note="単位は百万円から億円へ換算して記載している")]},
      forbid=["A15"])
check("A16 単語1つの'期'ではWARNを黙らせられない", "a",
      {"ledger": [ev(quote="当連結会計年度の売上高は1,234,567百万円となりました。",
                    value={"number": 1234567, "unit": "百万円", "period": "24年3月期"},
                    note="期")]},
      expect=["A16"])
check("A17 単語1つの'出所'ではWARNを黙らせられない", "a",
      {"sources": [src(source_class="T2")],
       "ledger": [ev(tags=["市場環境"], note="出所")]},
      expect=["A17"])
check("A17 実質的な説明があれば通る", "a",
      {"sources": [src(source_class="T2")],
       "ledger": [ev(tags=["市場環境"], note="出所は会社推計。第三者統計に代替なし")]},
      forbid=["A17"])

# =====================================================================
# A22: 改竄フラグの強制（capture_source.pyが検出→check_aが読む）
# =====================================================================
check("A22 raw_sha256_previousがあればFAIL", "a",
      {"sources": [src(raw_sha256_previous="deadbeef" * 8)]}, expect=["A22"])
check("A22 text_sha256_tamperedがあればFAIL", "a",
      {"sources": [src(text_sha256_tampered="deadbeef" * 8)]}, expect=["A22"])
check("A22 フラグが無ければ黙る", "a", {}, forbid=["A22"])

# =====================================================================
# A23〜A26: 原典再アクセス（reverify_log.jsonl との整合）
# =====================================================================
def build_reverify_log(tmp, rows):
    with open(os.path.join(tmp, "reverify_log.jsonl"), "w", encoding="utf-8") as fh:
        for r in rows:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")


def check_reverify(name, scope_body, reverify_rows, ledger=None, sources_extra=None,
                   expect=(), forbid=()):
    tmp = tempfile.mkdtemp()
    try:
        fixture = {"ledger": ledger} if ledger is not None else {}
        if sources_extra is not None:
            fixture["sources"] = sources_extra
        build(tmp, **fixture)
        with open(os.path.join(tmp, "scope.md"), "w", encoding="utf-8") as fh:
            fh.write(scope_body)
        build_reverify_log(tmp, reverify_rows)
        res = check_a.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            print(f"[{'PASS' if ok else 'FAIL'}] {name}")
            if missing:
                print(f"       未検出: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


CRIT_SCOPE = ("## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
             "|---|---|---|---|---|\n| N-01 | 売上高 | C | 有報 | 取得済(E-0001) |\n")

check_reverify("A23 Critical証拠だが再アクセス記録が無い",
               CRIT_SCOPE, [], expect=["A23"])
check_reverify("A23 再アクセス記録(match)があれば黙る",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "match",
                 "reverified_at": "2025-06-20T12:00:00+09:00"}],
               forbid=["A23"])
check_reverify("A24 hallucination判定なのに台帳に残っている",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "hallucination",
                 "detail": "URLが死んでいた", "reverified_at": "2025-06-20T12:00:00+09:00"}],
               expect=["A24"])
check_reverify("A25 context_reversedだがnoteに例外規定が無い",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "context_reversed",
                 "detail": "前後の条件で意味が反転", "reverified_at": "2025-06-20T12:00:00+09:00"}],
               expect=["A25"])
check_reverify("A25 noteに例外規定を実質的に書けば黙る",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "context_reversed",
                 "detail": "前後の条件で意味が反転", "reverified_at": "2025-06-20T12:00:00+09:00"}],
               ledger=[ev(note="低圧限定という条件付きの記述であり、例外規定として明記する")],
               forbid=["A25"])
check_reverify("A26 再アクセスが収集から48時間以上遅い",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "match",
                 "reverified_at": "2025-06-25T12:00:00+09:00"}],
               sources_extra=[src(retrieved_at="2025-06-20T12:00:00+09:00")],
               expect=["A26"])
check_reverify("A26 収集直後の再アクセスなら黙る",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "match",
                 "reverified_at": "2025-06-20T13:00:00+09:00"}],
               sources_extra=[src(retrieved_at="2025-06-20T12:00:00+09:00")],
               forbid=["A26"])
check_reverify("unreachableは進行を止めない(A24/A25発火せず)",
               CRIT_SCOPE,
               [{"eid": "E-0001", "url": "https://example.com", "outcome": "unreachable",
                 "detail": "アクセスできず", "reverified_at": "2025-06-20T12:00:00+09:00"}],
               forbid=["A23", "A24", "A25"])
check_reverify("A23 Normal項目には適用されない", "## 必要証拠リスト\n\n"
               "| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n|---|---|---|---|---|\n"
               "| N-01 | 従業員数 | N | 有報 | 取得済(E-0001) |\n",
               [], forbid=["A23"])

# =====================================================================
# ステータスLint
# =====================================================================
NEG_EV = ev(eid="E-0002", claim_type="negative",
            quote="なお、セグメント別の設備投資額については記載しておりません。",
            fact="有報を確認したが記載なし", tags=["財務.健全性"])

# =====================================================================
# S13: 存在しないE-IDによる[非公表]バイパスの防止（重大な抜け穴の修正）
# =====================================================================
check("S13 存在しないE-IDで[非公表]を偽装", "s",
      {"ledger": [ev()],
       "report": "設備投資 [非公表|確認:E-9999|代替:按分による推計を用いる]。\n"},
      expect=["S13"], forbid=["S01", "S12"])
check("S13 実在するnegative証拠なら黙る", "s",
      {"ledger": [ev(), NEG_EV],
       "report": "設備投資 [非公表|確認:E-0002|代替:按分による推計を用いる]。\n"},
      forbid=["S13", "S01", "S12"])

check("S01 非公表に確認E-IDが無い", "s",
      {"ledger": [ev()], "report": "セグメント別設備投資 [非公表|代替:按分]。\n"},
      expect=["S01"])
check("S12 非公表の確認証拠がnegativeでない", "s",
      {"ledger": [ev()], "report": "設備投資 [非公表|確認:E-0001|代替:按分]。\n"},
      expect=["S12"])
check("S12 negative証拠なら通る", "s",
      {"ledger": [ev(), NEG_EV], "report": "設備投資 [非公表|確認:E-0002|代替:按分]。\n"},
      forbid=["S01", "S12"])
check("S03/S04 未調査に次アクション・優先度が無い", "s",
      {"report": "顧客集中度 [未調査]。\n"}, expect=["S03", "S04"])
check("S05/S06 調査不可に障壁コード・依頼IDが無い", "s",
      {"report": "目標株価 [調査不可]。\n"}, expect=["S05", "S06"])
check("S06 依頼IDがあれば通る", "s",
      {"report": "目標株価 [調査不可|PRICED_DB|依頼:R-001]。\n"},
      forbid=["S05", "S06", "S11"])
check("S11 存在しない依頼ID", "s",
      {"report": "目標株価 [調査不可|PRICED_DB|依頼:R-999]。\n"}, expect=["S11"])
check("S07/S08/S09 推定の要件欠落", "s",
      {"report": "粗利率は42〜44%となる [推定]。\n"}, expect=["S07", "S08", "S09"])
check("S07-09 要件を満たせば通る", "s",
      {"report": "粗利率は42〜44%となる "
                 "[推定|根拠:E-0001|ロジック:外挿|確信度:中|反証:単価下落]。\n"},
      forbid=["S07", "S08", "S09"])
check("S10 曖昧語（不明）", "s", {"report": "市場シェアは不明である。\n"}, expect=["S10"])
check("S10 曖昧語（判然としない）", "s",
      {"report": "市場動向は判然としない。\n"}, expect=["S10"])

# =====================================================================
# S14 / 飽和判定: 「未特定」トークンと探索ログの整合
# =====================================================================
def build_search_log(tmp, rows):
    write_jsonl_helper = os.path.join(tmp, "search_log.jsonl")
    with open(write_jsonl_helper, "w", encoding="utf-8") as fh:
        for r in rows:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")


def check_s14(name, log_rows, report_line, expect=(), forbid=()):
    tmp = tempfile.mkdtemp()
    try:
        build(tmp, report=report_line)
        build_search_log(tmp, log_rows)
        res = lint_status.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            print(f"[{'PASS' if ok else 'FAIL'}] {name}")
            if missing:
                print(f"       未検出: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


SATURATED_LOG = [
    {"item": "N-01", "route": "T1", "query": "有報 セグメント別設備投資",
     "timestamp": "2026-08-01T10:00:00", "new_subjects": 2, "new_facts": 1},
    {"item": "N-01", "route": "T1", "query": "有報 設備の状況",
     "timestamp": "2026-08-02T10:00:00", "new_subjects": 0, "new_facts": 0},
    {"item": "N-01", "route": "T2", "query": "決算説明資料 設備投資 セグメント",
     "timestamp": "2026-08-03T10:00:00", "new_subjects": 0, "new_facts": 0},
    {"item": "N-01", "route": "T3", "query": "経済産業省 設備投資統計 業界別",
     "timestamp": "2026-08-04T10:00:00", "new_subjects": 0, "new_facts": 0},
]
check_s14("S14 飽和条件を満たせば通る",
          SATURATED_LOG, "設備投資額 [未特定|対象:N-01]。\n",
          forbid=["S14"])
check_s14("S14 ログが3件未満なら失敗", SATURATED_LOG[:2],
          "設備投資額 [未特定|対象:N-01]。\n", expect=["S14"])
check_s14("S14 直近3件の経路が重複していれば失敗",
          [dict(r, route="T1") for r in SATURATED_LOG],
          "設備投資額 [未特定|対象:N-01]。\n", expect=["S14"])
DUP_QUERY_LOG = [dict(r) for r in SATURATED_LOG]
DUP_QUERY_LOG[-1]["query"] = SATURATED_LOG[-2]["query"]  # 直近2件のクエリが同一
check_s14("S14 直近3件のクエリが重複していれば失敗",
          DUP_QUERY_LOG, "設備投資額 [未特定|対象:N-01]。\n", expect=["S14"])
NONZERO_LOG = [dict(r) for r in SATURATED_LOG]
NONZERO_LOG[-1]["new_facts"] = 1
check_s14("S14 直近3件に新規事実ありなら失敗",
          NONZERO_LOG, "設備投資額 [未特定|対象:N-01]。\n", expect=["S14"])
check_s14("S14 項目IDが無ければ失敗", SATURATED_LOG, "設備投資額 [未特定]。\n",
          expect=["S14"])

# =====================================================================
# L01/L02: 探索ログの多様性（V37/V38相当）
# =====================================================================
def check_l_test(name, log_rows, expect=(), forbid=()):
    tmp = tempfile.mkdtemp()
    try:
        build_search_log(tmp, log_rows)
        res = check_l.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            print(f"[{'PASS' if ok else 'FAIL'}] {name}")
            if missing:
                print(f"       未検出: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


check_l_test("L01 経路は違うがクエリが完全一致(V38)",
             [{"item": "N-01", "route": "T1", "query": "同一クエリ",
               "timestamp": "2026-08-01T10:00:00", "new_subjects": 0, "new_facts": 0},
              {"item": "N-01", "route": "T3", "query": "同一クエリ",
               "timestamp": "2026-08-02T10:00:00", "new_subjects": 0, "new_facts": 0}],
             expect=["L01"])
check_l_test("L01 経路もクエリも違えば発火しない", SATURATED_LOG, forbid=["L01"])
check_l_test("L02 独立再検証が収集時と同一クエリ(V37)",
             [{"item": "N-01", "route": "T1", "query": "有報 設備投資", "stage": "collection",
               "timestamp": "2026-08-01T10:00:00", "new_subjects": 1, "new_facts": 1},
              {"item": "N-01", "route": "T1", "query": "有報 設備投資", "stage": "reverify",
               "timestamp": "2026-08-10T10:00:00", "new_subjects": 0, "new_facts": 0}],
             expect=["L02"])
check_l_test("L02 再検証が異なるクエリなら発火しない",
             [{"item": "N-01", "route": "T1", "query": "有報 設備投資", "stage": "collection",
               "timestamp": "2026-08-01T10:00:00", "new_subjects": 1, "new_facts": 1},
              {"item": "N-01", "route": "T2", "query": "決算説明資料 設備投資", "stage": "reverify",
               "timestamp": "2026-08-10T10:00:00", "new_subjects": 0, "new_facts": 0}],
             forbid=["L02"])

# =====================================================================
# チェックB
# =====================================================================
OK_REPORT = ("# レポート\n\n産業機器セグメントの売上高は700,000百万円である [E-0001]。\n")
FULL_LEDGER = [
    ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期"}),
    ev(eid="E-0002", quote="産業機器セグメントの出荷台数は12.0百万台となりました。",
       fact="出荷台数12.0百万台",
       value={"number": 12.0, "unit": "百万台", "period": "25年3月期"}),
    ev(eid="E-0003", quote="前連結会計年度の産業機器セグメントの売上高は650,000百万円でありました。",
       fact="前期 産業機器売上 650,000百万円"),
    ev(eid="E-0004", claim_type="forecast",
       quote="次期の売上高は1,300,000百万円を見込んでおります。",
       fact="会社予想 売上高1,300,000百万円"),
    ev(eid="E-0005", quote="当社は上位顧客への依存度が高く、業績が変動する可能性があります。",
       fact="顧客集中リスクの記載", tags=["リスク"]),
]

check("B01 存在しないE-IDを参照", "b",
      {"report": "売上は700,000百万円 [E-9999]。\n", "figures": FIG_OK},
      expect=["B01"])
check("B02 出典の無い数値記述", "b",
      {"report": "# R\n\n海外売上比率は65%に達している。\n", "figures": FIG_OK},
      expect=["B02"])
check("B02 引用ブロックに隠した断定も検出", "b",
      {"report": "# R\n\n> 当社の競争優位は盤石であり首位である。\n", "figures": FIG_OK},
      expect=["B02"])
check("B02 短い断定も検出", "b",
      {"report": "# R\n\n同社は極めて優れた経営をしている。\n", "figures": FIG_OK},
      expect=["B02"])
check("B02 出典付きなら黙る", "b",
      {"report": OK_REPORT, "figures": FIG_OK}, forbid=["B02"])
check("B03 引用に無い数値", "b",
      {"report": "# R\n\nデバイス売上は450,000百万円である [E-0001]。\n",
       "figures": FIG_OK}, expect=["B03"])
check("B03 一致する数値は黙る", "b",
      {"report": OK_REPORT, "figures": FIG_OK}, forbid=["B03"])
check("B04 計算値の不一致", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n単価は99,999円/台 [計算:E-0001,E-0002|式:E-0001/E-0002]。\n",
       "figures": FIG_OK}, expect=["B04"])
check("B04 正しい計算は黙る", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n単価は58,333円/台 [計算:E-0001,E-0002|式:E-0001/E-0002]。\n",
       "figures": FIG_OK}, forbid=["B04"])
check("B05 推定の要件欠落", "b",
      {"report": "# R\n\n粗利率は42〜44%である [推定|根拠:E-0001|確信度:中]。\n",
       "figures": FIG_OK}, expect=["B05"])
check("B06 図の根拠列が空", "b",
      {"report": OK_REPORT,
       "figures": dict(FIG_OK, **{"F-01": "# F\n\n## データ表\n| 項目 | 値 | 根拠 |\n"
                                           "|---|---|---|\n| 売上 | 700000 |  |\n"})},
      expect=["B06"])
check("B07 使われていない証拠（取りこぼし）", "b",
      {"ledger": FULL_LEDGER, "report": OK_REPORT, "figures": FIG_OK},
      expect=["B07"])
check("B08 カテゴリの欠落", "b",
      {"report": OK_REPORT, "figures": FIG_OK}, expect=["B08"])
check("B08 見出しに名前を並べるだけでは黙らない", "b",
      {"report": OK_REPORT + "\n## 市場環境 競争環境 技術 ガバナンス\n",
       "figures": FIG_OK}, expect=["B08"])
check("B09 会社予想を断定で書く", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n次期売上高は1,300,000百万円となる [E-0004]。\n",
       "figures": FIG_OK}, expect=["B09"])
check("B09 予想と明示すれば黙る", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n会社は次期売上高を1,300,000百万円と予想している [E-0004]。\n",
       "figures": FIG_OK}, forbid=["B09"])
check("B13 引用と逆向きの否定", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n顧客集中リスクは存在しない [E-0005]。\n",
       "figures": FIG_OK}, expect=["B13"])
check("B13 negative証拠を使えば黙る", "b",
      {"ledger": FULL_LEDGER + [NEG_EV],
       "report": "# R\n\nセグメント別設備投資の開示は存在しない [E-0002]。\n"
                 .replace("E-0002", "E-0002"),
       "figures": FIG_OK}, forbid=[])
check("B14 単一時点で変化を主張", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n産業機器の売上は減少した [E-0003]。\n",
       "figures": FIG_OK}, expect=["B14"])
check("B14 2期分あれば黙る", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n産業機器の売上は増加した [E-0001][E-0003]。\n",
       "figures": FIG_OK}, forbid=["B14"])
check("B15 最小図解セットの不足", "b",
      {"report": OK_REPORT, "figures": {"F-01": FIG_OK["F-01"]}},
      expect=["B15"])
check("B15 F-01〜F-05が揃えば黙る", "b",
      {"report": OK_REPORT, "figures": FIG_OK}, forbid=["B15"])
check("B16 引用の取り違え（語が別の箇所にある）", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\n出荷台数の増減 [E-0001]。\n".replace("の増減", "は重要である"),
       "figures": FIG_OK}, expect=["B16"])
check("B17 主語のすり替え", "b",
      {"ledger": FULL_LEDGER,
       "report": "# R\n\nデバイスセグメントの売上高は700,000百万円である [E-0001]。\n",
       "figures": FIG_OK}, expect=["B17"])
check("メタ行（ゲート証跡）は主張として扱わない", "b",
      {"report": "# R\n\n<!-- gate:B PASS receipt=abc123 -->\n" + OK_REPORT,
       "figures": FIG_OK}, forbid=["B02"])


# =====================================================================
# B21: 最上級・排他の主張の根拠要求（既存語彙の用途拡張）
# =====================================================================
SUPER_RAW = (BASE_RAW +
             "当社の産業機器セグメントは国内シェア45%で第1位であります。\n")
SUPER_LEDGER = [
    ev(),  # 自社の売上高だけの引用（順位・シェアの根拠は無い）
    ev(eid="E-0002",
       anchor_head="当社の産業機器セグメントは国内シェア",
       anchor_tail="第1位であります。",
       fact="国内シェア45%で第1位",
       value={"number": 45, "unit": "%", "period": "25年3月期"}),
]
check("B21 最上級だが引用に順位・シェアの根拠が無い", "b",
      {"raw": {"S-001": SUPER_RAW}, "ledger": SUPER_LEDGER,
       "report": "# R\n\n当社は業界最大の産業機器メーカーである [E-0001]。\n",
       "figures": FIG_OK}, expect=["B21"])
check("B21 順位の証拠があれば通る", "b",
      {"raw": {"S-001": SUPER_RAW}, "ledger": SUPER_LEDGER,
       "report": "# R\n\n当社は国内首位の産業機器メーカーである [E-0002]。\n",
       "figures": FIG_OK}, forbid=["B21"])
check("B21 [推定]に落とせば対象外", "b",
      {"raw": {"S-001": SUPER_RAW}, "ledger": SUPER_LEDGER,
       "report": "# R\n\n業界最大とみられる "
                 "[推定|根拠:E-0001|ロジック:売上規模から|確信度:低|反証:他社開示]。\n",
       "figures": FIG_OK}, forbid=["B21"])
check("B21 最上級を含まない主張には発火しない", "b",
      {"raw": {"S-001": SUPER_RAW}, "ledger": SUPER_LEDGER,
       "report": OK_REPORT, "figures": FIG_OK},
      forbid=["B21"])

# =====================================================================
# B20: カテゴリ一括スキップの可視化
# =====================================================================
check("B20 複数カテゴリを1つの理由で一括スキップ", "b",
      {"report": OK_REPORT + "\n市場環境 / 競争環境 / 技術 / ガバナンス "
               "[未調査|意図的|理由:スコープ外|次:次回|優先:L]。\n",
       "figures": FIG_OK}, expect=["B20"])
check("B20 単独カテゴリなら発火しない", "b",
      {"report": OK_REPORT + "\n市場環境 [未調査|意図的|理由:スコープ外|次:次回|優先:L]。\n",
       "figures": FIG_OK}, forbid=["B20"])

# =====================================================================
# B19: 主体（subject）の一致 — 統制語彙による取り違え検出
# =====================================================================
SUBJECT_LEDGER = [
    ev(subject="産業機器セグメント"),
    ev(eid="E-0002", subject="デバイスセグメント",
       quote="デバイスセグメントの売上高は400,000百万円であります。",
       fact="デバイスセグメント売上400,000百万円",
       value={"number": 400000, "unit": "百万円", "period": "25年3月期"}),
]
check("B19 主体の取り違え（産業機器→デバイスと誤記）", "b",
      {"ledger": SUBJECT_LEDGER,
       "report": "# R\n\nデバイスセグメントの売上高は700,000百万円である [E-0001]。\n",
       "figures": FIG_OK}, expect=["B19"])
check("B19 主体が一致していれば黙る", "b",
      {"ledger": SUBJECT_LEDGER, "report": OK_REPORT, "figures": FIG_OK},
      forbid=["B19"])
check("B19 subject未設定の証拠には適用されない", "b",
      {"report": "デバイスセグメントの売上高は700,000百万円である [E-0001]。\n",
       "figures": FIG_OK}, forbid=["B19"])

# =====================================================================
# A17: ソース選択の妥当性
# =====================================================================
check("A17 実績値をT5ソースで取っている", "a",
      {"sources": [src(source_class="T5")],
       "ledger": [ev(confidence="中", tags=["財務.実績"])]},
      expect=["A17"])
check("A17 実績値がT1なら黙る", "a",
      {"ledger": [ev(tags=["財務.実績"])]}, forbid=["A17"])
check("A17 市場規模を発行体資料(T2)で取っている", "a",
      {"sources": [src(source_class="T2")],
       "ledger": [ev(tags=["市場環境"])]}, expect=["A17"])
check("A17 市場規模がT3公的統計なら黙る", "a",
      {"sources": [src(source_class="T3")],
       "ledger": [ev(tags=["市場環境"], confidence="中")]}, forbid=["A17"])
check("A17 noteに出所と限界を書けば許容", "a",
      {"sources": [src(source_class="T2")],
       "ledger": [ev(tags=["市場環境"], note="出所は会社推計。第三者統計に代替なし")]},
      forbid=["A17"])

# =====================================================================
# A20: 連結/単体・累計/単四半期の一致
# =====================================================================
check("A20 単体と明記した数値が引用は連結", "a",
      {"raw": {"S-001": "連結会計年度における産業機器セグメントの売上高は"
                        "700,000百万円であります。"},
       "ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期",
                            "basis_scope": "単体"},
                    quote="連結会計年度における産業機器セグメントの売上高は"
                          "700,000百万円であります。")]},
      expect=["A20"])
check("A20 scope一致なら黙る", "a",
      {"raw": {"S-001": "連結会計年度における産業機器セグメントの売上高は"
                        "700,000百万円であります。"},
       "ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期",
                            "basis_scope": "連結"},
                    quote="連結会計年度における産業機器セグメントの売上高は"
                          "700,000百万円であります。")]},
      forbid=["A20"])
check("A20 scope未指定なら対象外", "a",
      {"ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期"})]},
      forbid=["A20"])
check("A20 原文のどこにも区分の記載が無ければWARN", "a",
      {"raw": {"S-001": "産業機器セグメントの売上高は700,000百万円であります。"},
       "ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期",
                            "basis_scope": "連結"})]},
      expect=["A20"])
check("A20 表ヘッダ等の近傍に区分があれば黙る（実データ対応）", "a",
      {"raw": {"S-001": "（単位：百万円）連結ベースの数値であります。\n"
                        "産業機器セグメントの売上高は700,000百万円であります。"},
       "ledger": [ev(value={"number": 700000, "unit": "百万円", "period": "25年3月期",
                            "basis_scope": "連結"})]},
      forbid=["A20"])

# =====================================================================
# A21: 複数期の要約表からの引用
# =====================================================================
A21_RAW = "25年3月期の売上高は1,234,567百万円、24年3月期は1,100,000百万円でした。"
check("A21 複数期が並ぶ引用でconfidence高", "a",
      {"raw": {"S-001": A21_RAW},
       "ledger": [ev(quote=A21_RAW,
                    value={"number": 1234567, "unit": "百万円", "period": "25年3月期"})]},
      expect=["A21"])
check("A21 単一期の引用は対象外", "a", {}, forbid=["A21"])
check("A21 要約表由来と明記すれば黙る", "a",
      {"raw": {"S-001": A21_RAW},
       "ledger": [ev(quote=A21_RAW,
                    value={"number": 1234567, "unit": "百万円", "period": "25年3月期"},
                    note="決算ハイライト(5期推移)から抽出")]},
      forbid=["A21"])

# =====================================================================
# A19: 文書全体の前提ゾーン検査（冒頭の前提 vs 後半の根拠）
# =====================================================================
LONG_DOC_WITH_PREMISE = (
    "算定の基礎: 本資料の数値は為替レート1ドル150円、原材料価格を前年並みと"
    "仮定して算定しております。" + "本文の記述です。" * 200 +
    "産業機器セグメントの売上高は700,000百万円であります。" +
    "本文の記述です。" * 50
)
check("A19 前提条件が冒頭、根拠が離れた位置", "a",
      {"raw": {"S-001": LONG_DOC_WITH_PREMISE}, "ledger": [ev()]},
      expect=["A19"])
check("A19 noteに前提確認済みと書けば黙る", "a",
      {"raw": {"S-001": LONG_DOC_WITH_PREMISE},
       "ledger": [ev(note="算定の基礎（為替150円）を確認済み、矛盾なし")]},
      forbid=["A19"])
check("A19 前提の記載が無い長文書では発火しない", "a",
      {"raw": {"S-001": "本文の記述です。" * 200 +
              "産業機器セグメントの売上高は700,000百万円であります。" +
              "本文の記述です。" * 50},
       "ledger": [ev()]}, forbid=["A19"])
check("A19 引用自体が前提ゾーン内なら発火しない", "a",
      {"raw": {"S-001": "算定の基礎: 為替は150円と仮定。"
              "産業機器セグメントの売上高は700,000百万円であります。" +
              "本文の記述です。" * 200},
       "ledger": [ev()]}, forbid=["A19"])
check("A19 短い通常の文書では発火しない（回帰）", "a", {}, forbid=["A19"])

# =====================================================================
# A18: 引用の切り出し範囲（文脈）
# =====================================================================
HEDGE_RAW = ("当社の主力製品は堅調に推移しました。"
             "ただし、為替の影響を除くと売上高は減少しております。"
             "産業機器セグメントの売上高は700,000百万円であります。")
check("A18 直後の留保を落として引用している", "a",
      {"raw": {"S-001": "産業機器セグメントの売上高は700,000百万円であります。"
                        "ただし、これには一過性の要因が含まれます。"},
       "ledger": [ev(anchor_head="産業機器セグメントの売上高は",
                    anchor_tail="であります。")]}, expect=["A18"])
check("A18 文の途中で切っている", "a",
      {"ledger": [ev(anchor_head="産業機器セグメントの売上高は",
                    anchor_tail="700,000百万円で")]},
      expect=["A18"])
check("A18 文頭から文末まで、前後に留保が無ければ黙る", "a",
      {"raw": {"S-001": "産業機器セグメントの売上高は700,000百万円であります。\n"
                        "従業員数は5,000名です。"},
       "ledger": [ev()]}, forbid=["A18"])

# =====================================================================
# B18: 採用した問いの網羅
# =====================================================================
def with_scope(fixture, scope):
    return fixture, scope


tmp_scope = tempfile.mkdtemp()
try:
    build(tmp_scope, ledger=FULL_LEDGER, report=OK_REPORT, figures=FIG_OK)
    with open(os.path.join(tmp_scope, "scope.md"), "w", encoding="utf-8") as fh:
        fh.write("## 採用した問い\n| Q-R4 | 崩壊条件 |\n| Q-B1 | 誰が払うか |\n")
    res = check_b.run(tmp_scope)
    codes = {f.code for f in res.findings}
    ok = "B18" in codes
    RESULTS.append(("B18 採用した問いがレポートに無い", ok,
                    [] if ok else ["B18"], [], sorted(codes)))
    if not ok:
        print("[FAIL] B18 採用した問いがレポートに無い")
finally:
    shutil.rmtree(tmp_scope, ignore_errors=True)

tmp_scope2 = tempfile.mkdtemp()
try:
    build(tmp_scope2, ledger=FULL_LEDGER,
          report=OK_REPORT + "\n顧客集中度 [未調査|次:有報確認|優先:M] (Q-R4)。\n",
          figures=FIG_OK)
    with open(os.path.join(tmp_scope2, "scope.md"), "w", encoding="utf-8") as fh:
        fh.write("## 採用した問い\n| Q-R4 | 崩壊条件 |\n")
    res = check_b.run(tmp_scope2)
    codes = {f.code for f in res.findings}
    ok = "B18" not in codes
    RESULTS.append(("B18 問いIDを本文に残せば黙る", ok, [], [] if ok else ["B18"],
                    sorted(codes)))
    if not ok:
        print("[FAIL] B18 問いIDを本文に残せば黙る")
finally:
    shutil.rmtree(tmp_scope2, ignore_errors=True)

# =====================================================================
# 実データ形式の回帰テスト（開示文書の実際の書式で誤検出しないこと）
# =====================================================================
REAL_RAW = """2027年3月期 第1四半期決算短信〔日本基準〕（連結）

（セグメント情報等の注記）
（単位：百万円）
外部顧客への売上高 情報通信 264,402 エレクトロニクス 35,347 合計 402,009
セグメント利益 情報通信 97,959 エレクトロニクス 444 合計 104,828

（2）四半期連結損益計算書（単位：百万円）
売上高 267,908 / 402,009
受取手形、売掛金及び契約資産 252,623 / 314,525

3．2027年3月期の連結業績予想
通期 売上高 1,755,000百万円 48.4% 営業利益 432,000百万円 128.9%
"""
REAL_SRC = [src(period_context="2027年3月期第1四半期")]
check("実データ: 表ヘッダの単位でA15が誤検出しない", "a",
      {"raw": {"S-001": REAL_RAW}, "sources": REAL_SRC,
       "ledger": [ev(quote="外部顧客への売上高 情報通信 264,402 エレクトロニクス 35,347 合計 402,009",
                    fact="情報通信セグメント売上264,402百万円",
                    value={"number": 264402, "unit": "百万円",
                           "period": "2027年3月期第1四半期", "basis_scope": "連結"},
                    note="セグメント情報の当第1四半期表より")]},
      forbid=["A15", "A18", "A20"])
check("実データ: 業績予想表でA18が誤検出しない", "a",
      {"raw": {"S-001": REAL_RAW}, "sources": REAL_SRC,
       "ledger": [ev(claim_type="forecast",
                    quote="通期 売上高 1,755,000百万円 48.4% 営業利益 432,000百万円 128.9%",
                    fact="通期会社予想 売上高1,755,000百万円",
                    value={"number": 1755000, "unit": "百万円", "period": "2027年3月期"},
                    note="連結業績予想。会社予想であり実績ではない")]},
      forbid=["A18"])
check("実データ: 2期比較行でA18が誤検出しない", "a",
      {"raw": {"S-001": REAL_RAW}, "sources": REAL_SRC,
       "ledger": [ev(quote="売上高 267,908 / 402,009",
                    fact="全社売上高 前1Q267,908→当1Q402,009百万円",
                    value={"number": 402009, "unit": "百万円",
                           "period": "2027年3月期第1四半期"},
                    note="四半期連結損益計算書より。左が前年同期、右が当期")]},
      forbid=["A18", "A15"])

REAL_LEDGER = [
    ev(quote="外部顧客への売上高 情報通信 264,402 エレクトロニクス 35,347 合計 402,009",
       fact="情報通信セグメント売上264,402百万円",
       value={"number": 264402, "unit": "百万円", "period": "2027年3月期第1四半期"}),
]
check("実データ: メタ情報行でB02が誤検出しない", "b",
      {"ledger": REAL_LEDGER,
       "report": ("# R\n\n- 調査基準日: 2026-08-26 ／ 調査範囲: 2027年3月期第1四半期\n"
                  "- 使用ソース: T1 1本（決算短信）\n"),
       "figures": FIG_OK}, forbid=["B02"])
check("実データ: 根拠提示後の解釈文でB02が誤検出しない", "b",
      {"ledger": REAL_LEDGER,
       "report": ("# R\n\n情報通信の売上高は264,402百万円である [E-0001]。\n"
                  "ただしこれは裏返せば、単一セグメントへの依存が高いことを意味する。\n"),
       "figures": FIG_OK}, forbid=["B02"])
check("実データ: 付録節の記述でB02が誤検出しない", "b",
      {"ledger": REAL_LEDGER,
       "report": ("# R\n\n情報通信の売上高は264,402百万円である [E-0001]。\n\n"
                  "## 付録A. 本調査の限界\n\n"
                  "- 単一ソースのみに依拠しており、需要の持続性を判断する材料が不足している\n"),
       "figures": FIG_OK}, forbid=["B02"])
check("実データ: 節が変われば解釈文の免除は効かない（抜け穴防止）", "b",
      {"ledger": REAL_LEDGER,
       "report": ("# R\n\n情報通信の売上高は264,402百万円である [E-0001]。\n\n"
                  "## 2. 競争環境\n\n"
                  "つまり当社は業界で圧倒的な地位を確立しているといえる。\n"),
       "figures": FIG_OK}, expect=["B02"])

# =====================================================================
# 静的検査: スクリプトが外部通信をしないこと
# =====================================================================
FORBIDDEN = ["urllib.request", "import requests", "http.client", "socket",
             "api.anthropic.com", "ANTHROPIC_API_KEY", "subprocess"]
offenders = []
for fn in sorted(os.listdir(os.path.join(SKILL, "scripts"))):
    if not fn.endswith(".py"):
        continue
    body = open(os.path.join(SKILL, "scripts", fn), encoding="utf-8").read()
    for token in FORBIDDEN:
        if token in body:
            offenders.append(f"{fn}: {token}")
RESULTS.append(("静的検査: スクリプトに外部通信・課金経路が無い",
                not offenders, offenders, [], []))
if offenders:
    print("[FAIL] 静的検査: 外部通信の可能性")
    for o in offenders:
        print(f"       {o}")

# =====================================================================
# B22: 必要証拠リストのCritical項目の充足検証
# =====================================================================
def check_b22(name, scope_body, expect=(), forbid=()):
    tmp = tempfile.mkdtemp()
    try:
        build(tmp, ledger=FULL_LEDGER, report=OK_REPORT, figures=FIG_OK)
        with open(os.path.join(tmp, "scope.md"), "w", encoding="utf-8") as fh:
            fh.write(scope_body)
        res = check_b.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            print(f"[{'PASS' if ok else 'FAIL'}] {name}")
            if missing:
                print(f"       未検出: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


check_b22("B22 Critical項目が未取得のまま",
          "## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
          "|---|---|---|---|---|\n| N-01 | セグメント別設備投資 | C | 有報 | 未取得 |\n",
          expect=["B22"])
check_b22("B22 E-IDで解決済みなら黙る",
          "## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
          "|---|---|---|---|---|\n| N-01 | 売上高 | C | 有報 | 取得済(E-0001) |\n",
          forbid=["B22"])
check_b22("B22 非公表と判断済みなら黙る",
          "## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
          "|---|---|---|---|---|\n| N-01 | セグメント別設備投資 | C | 有報 | 非公表 |\n",
          forbid=["B22"])
check_b22("B22 Normal項目の未取得は対象外",
          "## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
          "|---|---|---|---|---|\n| N-01 | 従業員の平均年齢 | N | 有報 | 未取得 |\n",
          forbid=["B22"])
check_b22("B22 未特定と判断済みなら黙る",
          "## 必要証拠リスト\n\n| ID | 証拠要件 | 重要度 | 想定ソース | 状態 |\n"
          "|---|---|---|---|---|\n| N-01 | セグメント別設備投資 | C | 有報 | 未特定(N-01) |\n",
          forbid=["B22"])

# =====================================================================
# 未整備だったコードの補完テスト（B10 / S02 / L00 / C04）
# =====================================================================
LONG_QUOTE_EV = ev(quote="産業機器セグメントの売上高は700,000百万円であります。",
                   fact="産業機器セグメント売上700,000百万円")
check("B10 同一ソースから複数回の直接引用", "b",
      {"ledger": [LONG_QUOTE_EV],
       "report": ("# R\n\n「産業機器セグメントの売上高は700,000百万円であります」"
                  " [E-0001]。\n"
                  "再度「産業機器セグメントの売上高は700,000百万円であります」"
                  " [E-0001]。\n"),
       "figures": FIG_OK}, expect=["B10"])
check("S02 [非公表]の代替アプローチが空欄に近い", "s",
      {"ledger": [ev(), NEG_EV],
       "report": "設備投資 [非公表|確認:E-0002|代替:x]。\n"}, expect=["S02"])
check("S02 代替が具体的なら黙る", "s",
      {"ledger": [ev(), NEG_EV],
       "report": "設備投資 [非公表|確認:E-0002|代替:全社CAPEXを売上比で按分する]。\n"},
      forbid=["S02"])
check_l_test("L00 探索ログが空ならWARN", [], expect=["L00"])

# =====================================================================
# ゲートC: 独立レビュー(GC)の実施確認
# =====================================================================
import check_c_claims_review as check_c  # noqa: E402


def check_gate_c(name, report, claims_review_content, expect=(), forbid=()):
    tmp = tempfile.mkdtemp()
    try:
        build(tmp, ledger=FULL_LEDGER, report=report, figures=FIG_OK)
        if claims_review_content is not None:
            with open(os.path.join(tmp, "checks", "claims_review.md"),
                      "w", encoding="utf-8") as fh:
                fh.write(claims_review_content)
        res = check_c.run(tmp)
        codes = {f.code for f in res.findings}
        missing = [c for c in expect if c not in codes]
        extra = [c for c in forbid if c in codes]
        ok = not missing and not extra
        RESULTS.append((name, ok, missing, extra, sorted(codes)))
        if VERBOSE or not ok:
            mark = "PASS" if ok else "FAIL"
            print(f"[{mark}] {name}")
            if missing:
                print(f"       検出されるべきだが出なかった: {missing}")
            if extra:
                print(f"       誤検出: {extra}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


MANY_CLAIMS_REPORT = ("# R\n\n" + "\n".join(
    "産業機器セグメントの出荷台数は12.0百万台となりました [E-0002]。"
    for _ in range(6)))

REVIEWED_TABLE_OK = (
    "| # | 箇所 | 主張 | 根拠 | 種別 | 引用 | 判定 | コメント |\n"
    "|---|---|---|---|---|---|---|---|\n" +
    "\n".join(f"| {i} | report.md:L1 | 主張{i} | E-0001 | fact | 引用{i} | 支持 |  |"
              for i in range(1, 7)))
REVIEWED_TABLE_UNRESOLVED = REVIEWED_TABLE_OK.replace(
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 支持 |  |",
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 不支持 | 主語が違う |")
REVIEWED_TABLE_EMPTY = REVIEWED_TABLE_OK.replace(
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 支持 |  |",
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 |  |  |")

check_gate_c("C00 大規模なのにclaims_review.mdが未生成",
             MANY_CLAIMS_REPORT, None, expect=["C00"])
check_gate_c("C01 判定が空欄の行がある",
             MANY_CLAIMS_REPORT, REVIEWED_TABLE_EMPTY, expect=["C01"])
check_gate_c("C02 不支持が残っている",
             MANY_CLAIMS_REPORT, REVIEWED_TABLE_UNRESOLVED, expect=["C02"])
check_gate_c("C00-C03 全件支持ならPASS",
             MANY_CLAIMS_REPORT, REVIEWED_TABLE_OK,
             forbid=["C00", "C01", "C02", "C03"])
check_gate_c("C04 表の行数がレポートの主張数より少ない（再抽出漏れ）",
             MANY_CLAIMS_REPORT,
             ("| # | 箇所 | 主張 | 根拠 | 種別 | 引用 | 判定 | コメント |\n"
              "|---|---|---|---|---|---|---|---|\n"
              "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 支持 | 確認済 |\n"),
             expect=["C04"])
check_gate_c("C00 主張が無ければWARNのみ（対象0件）",
             "# R\n\n本文のみ。\n", None, expect=["C00"])

# --- C05/C06: ラバースタンプ（機械的な支持連発）の兆候検出 ---
MANY_CLAIMS_REPORT2 = ("# R\n\n" + "\n".join(
    f"産業機器セグメントの出荷台数は12.0百万台となりました[{i}] [E-0002]。"
    for i in range(6)))
ALL_SUPPORT_SAME_COMMENT = (
    "| # | 箇所 | 主張 | 根拠 | 種別 | 引用 | 判定 | コメント |\n"
    "|---|---|---|---|---|---|---|---|\n" +
    "\n".join(f"| {i} | report.md:L1 | 主張{i} | E-0001 | fact | 引用{i} "
              "| 支持 | 問題なし |" for i in range(1, 7)))
check_gate_c("C05 全件『支持』かつ大量件数はラバースタンプの疑い",
             MANY_CLAIMS_REPORT2, ALL_SUPPORT_SAME_COMMENT, expect=["C05", "C06"])
MIXED_TABLE = (
    "| # | 箇所 | 主張 | 根拠 | 種別 | 引用 | 判定 | コメント |\n"
    "|---|---|---|---|---|---|---|---|\n"
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 支持 | 数値が一致 |\n"
    "| 2 | report.md:L1 | 主張2 | E-0001 | fact | 引用2 | 判断不能 | 引用がやや弱い |\n"
    "| 3 | report.md:L1 | 主張3 | E-0001 | fact | 引用3 | 支持 | セグメント名も一致 |\n"
    "| 4 | report.md:L1 | 主張4 | E-0001 | fact | 引用4 | 支持 | 期も確認した |\n"
    "| 5 | report.md:L1 | 主張5 | E-0001 | fact | 引用5 | 支持 | 単位も確認 |\n")
check_gate_c("C05/C06 多様な判定・コメントなら発火しない",
             MANY_CLAIMS_REPORT2, MIXED_TABLE, forbid=["C05", "C06"])

# =====================================================================
# export_wcheck: C-1/C-2/C-3 Wチェック結果の集約とExcel出力
# =====================================================================
import export_wcheck as wcheck  # noqa: E402


def check_wcheck(name, fixture, get_field, expect_value, reverify_rows=None, claims_md=None):
    tmp = tempfile.mkdtemp()
    try:
        build(tmp, **fixture)
        if reverify_rows is not None:
            with open(os.path.join(tmp, "reverify_log.jsonl"), "w", encoding="utf-8") as fh:
                for r in reverify_rows:
                    fh.write(json.dumps(r, ensure_ascii=False) + "\n")
        if claims_md is not None:
            os.makedirs(os.path.join(tmp, "checks"), exist_ok=True)
            with open(os.path.join(tmp, "checks", "claims_review.md"),
                      "w", encoding="utf-8") as fh:
                fh.write(claims_md)
        rows = wcheck.build_summary(tmp)
        actual = get_field(rows[0]) if rows else None
        ok = actual == expect_value
        RESULTS.append((name, ok, [] if ok else [str(expect_value)],
                        [] if ok else [str(actual)], []))
        if VERBOSE or not ok:
            print(f"[{'PASS' if ok else 'FAIL'}] {name}"
                 + ("" if ok else f" — expected={expect_value!r} actual={actual!r}"))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


WC_TABLE_UNSUPPORTED = (
    "| # | 箇所 | 主張 | 根拠 | 種別 | 引用 | 判定 | コメント |\n"
    "|---|---|---|---|---|---|---|---|\n"
    "| 1 | report.md:L1 | 主張1 | E-0001 | fact | 引用1 | 不支持 | 一致しない |\n")
WC_TABLE_SUPPORT = WC_TABLE_UNSUPPORTED.replace("不支持", "支持")

check_wcheck("wcheck: C-1がclaims_reviewの不支持を問題ありに変換",
             {"ledger": [ev()]}, lambda r: r["c1_status"], wcheck.ISSUE,
             claims_md=WC_TABLE_UNSUPPORTED)
check_wcheck("wcheck: C-1が支持を問題なしに変換",
             {"ledger": [ev()]}, lambda r: r["c1_status"], wcheck.NO_ISSUE,
             claims_md=WC_TABLE_SUPPORT)
check_wcheck("wcheck: C-2がhallucinationを問題ありに変換",
             {"ledger": [ev()]}, lambda r: r["c2_status"], wcheck.ISSUE,
             reverify_rows=[{"eid": "E-0001", "outcome": "hallucination", "detail": "URL404"}])
check_wcheck("wcheck: C-2記録が無ければ未実施",
             {"ledger": [ev()]}, lambda r: r["c2_status"], wcheck.NOT_DONE)
check_wcheck("wcheck: C-3が単位不一致を問題ありに変換",
             {"ledger": [ev(value={"number": 700000, "unit": "億円", "period": "25年3月期"})]},
             lambda r: r["c3_status"], wcheck.ISSUE)
check_wcheck("wcheck: C-3理由は機械生成メッセージそのまま(自由記述でない)",
             {"ledger": [ev(value={"number": 700000, "unit": "億円", "period": "25年3月期"})]},
             lambda r: "A15" in r["c3_reason"], True)
check_wcheck("wcheck: 清潔な証拠はC-3問題なし・C-1/C-2未実施",
             {"ledger": [ev()]},
             lambda r: (r["c3_status"], r["c1_status"], r["c2_status"]),
             (wcheck.NO_ISSUE, wcheck.NOT_DONE, wcheck.NOT_DONE))

# Excel出力の実ファイル検証（openpyxl未導入環境ではスキップ）
try:
    import openpyxl as _openpyxl_check
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

if _HAS_OPENPYXL:
    _tmp = tempfile.mkdtemp()
    try:
        build(_tmp, ledger=[ev()])
        _rows = wcheck.build_summary(_tmp)
        _out = os.path.join(_tmp, "checks", "out.xlsx")
        _path = wcheck.export_xlsx(_tmp, _rows, _out)
        _ok = _path is not None and os.path.exists(_out)
        if _ok:
            _wb = _openpyxl_check.load_workbook(_out)
            _ws = _wb.active
            _ok = _ws.cell(row=1, column=1).value == "EID" and _ws.max_row == 2
        RESULTS.append(("wcheck: Excelファイルが実際に生成されヘッダが正しい", _ok,
                        [] if _ok else ["xlsx検証失敗"], [], []))
        if not _ok:
            print("[FAIL] wcheck: Excelファイルが実際に生成されヘッダが正しい")
    finally:
        shutil.rmtree(_tmp, ignore_errors=True)
else:
    print("[SKIP] openpyxl未導入のためExcel実ファイル検証をスキップ")

total = len(RESULTS)
failed = [r for r in RESULTS if not r[1]]
print("\n" + "=" * 60)
print(f"網羅テスト: {total - len(failed)}/{total} 成功")
if failed:
    print("\n失敗:")
    for name, _, missing, extra, codes in failed:
        print(f"  - {name}")
        if missing:
            print(f"      未検出: {missing}")
        if extra:
            print(f"      誤検出: {extra}")
        print(f"      実際: {codes}")
sys.exit(1 if failed else 0)
